import os
import xlrd
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment

def converter_e_processar(arquivo_entrada, arquivo_saida, arquivo_destino):
    # Abrir o arquivo .xls para leitura
    wb_xls = xlrd.open_workbook(arquivo_entrada)
    ws_xls = wb_xls.sheet_by_index(0)

    # Criar um novo arquivo .xlsx
    wb_xlsx = Workbook()
    ws_xlsx = wb_xlsx.active

    # Copiar os dados do arquivo .xls para o arquivo .xlsx
    for row in range(ws_xls.nrows):
        for col in range(ws_xls.ncols):
            ws_xlsx.cell(row=row+1, column=col+1).value = ws_xls.cell_value(row, col)

    # Excluir as primeiras 7 linhas
    ws_xlsx.delete_rows(1, 7)
    
    # Salvar o arquivo .xlsx
    wb_xlsx.save(arquivo_saida)

    # Remover linhas em branco
    wb = load_workbook(arquivo_saida)
    ws = wb.active
    
    # Processar e adicionar nome do funcionário nas linhas de benefícios
    nome_funcionario = None
    for row in range(1, ws.max_row + 1):
        nome_na_coluna_C = ws.cell(row=row, column=3).value
        if nome_na_coluna_C and nome_na_coluna_C.strip():  # Verifica se há um nome de funcionário na terceira coluna
            nome_funcionario = nome_na_coluna_C.strip()
        if nome_funcionario and not nome_na_coluna_C:
            # Adiciona o nome do funcionário na linha atual
            ws.cell(row=row, column=3).value = nome_funcionario
    
    # Mesclar, centralizar e quebrar texto
    for row in ws.iter_rows():
        for cell in row:
            if cell.value:
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Remover colunas em branco
    colunas_para_remover = []
    for col in range(ws.max_column, 0, -1):
        if all(not ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)):
            colunas_para_remover.append(col)
    for col_index in colunas_para_remover:
        ws.delete_cols(col_index)

    # Remover linhas sem conteúdo na coluna I
    linhas_para_remover = []
    for row in range(1, ws.max_row + 1):
        if not ws.cell(row=row, column=9).value:  # Verifica a coluna I (9ª coluna)
            linhas_para_remover.append(row)

    for row_index in sorted(linhas_para_remover, reverse=True):
        ws.delete_rows(row_index)

    # Remover colunas em branco novamente
    colunas_para_remover = []
    for col in range(ws.max_column, 0, -1):
        if all(not ws.cell(row=row, column=col).value for row in range(1, ws.max_row + 1)):
            colunas_para_remover.append(col)
    for col_index in colunas_para_remover:
        ws.delete_cols(col_index)

    # Substituir ":" por "," na coluna F
    for row in range(1, ws.max_row + 1):
        valor_f = ws.cell(row=row, column=6).value
        if valor_f and ":" in valor_f:
            ws.cell(row=row, column=6).value = valor_f.replace(":", ",")

    # Salvar as alterações
    wb.save(arquivo_saida)

    # Verificar se o arquivo de destino existe
    if not os.path.exists(arquivo_destino):
        # Se não existir, criar um novo arquivo
        wb_destino = Workbook()
        wb_destino.save(arquivo_destino)

    # Abrir o arquivo de destino
    wb_destino = load_workbook(arquivo_destino)
    ws_destino = wb_destino.active
    
    # Obter a última linha da planilha de destino para começar a copiar os dados
    ultima_linha_destino = ws_destino.max_row

    # Copiar os valores da planilha de origem para a planilha de destino
    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            valor_celula_origem = ws.cell(row=row, column=col).value
            ws_destino.cell(row=ultima_linha_destino + row, column=col).value = valor_celula_origem

    # Salvar a planilha de destino com as alterações
    wb_destino.save(arquivo_destino)

# Caminho do arquivo de entrada .xls, do arquivo de saída .xlsx, e do arquivo de destino .xlsx
arquivo_entrada_xls = r"C:\Users\tiago\botao\Movimentos.xls"
arquivo_saida_xlsx = r"C:\Users\tiago\botao\Movimentos_processado.xlsx"
arquivo_destino_xlsx = r"C:\Users\tiago\botao\Movimentação_FLorence_Mensal.xlsx"

# Chamar a função para converter, processar e remover as linhas sem conteúdo na coluna I do arquivo .xls
converter_e_processar(arquivo_entrada_xls, arquivo_saida_xlsx, arquivo_destino_xlsx)
